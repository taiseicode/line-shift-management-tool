import hmac
import re
from urllib.parse import urlencode
from datetime import datetime, timedelta, date
from io import BytesIO

from flask import Blueprint, request, redirect, session, send_file, jsonify, make_response, render_template
from markupsafe import Markup
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from config import ADMIN_PASSWORD
from repositories.confirmed_shift_repository import (
    delete_confirmed_shift,
    get_confirmed_shift_by_id,
    get_confirmed_shift_decisions_by_date,
    get_confirmed_shift_decisions_range,
    get_confirmed_shifts_by_date,
    get_confirmed_shifts_range,
    upsert_confirmed_shift,
)
from repositories.required_staff_repository import upsert_required_staff, get_required_staff_range
from repositories.settings_repository import delete_setting, get_setting, get_settings, upsert_setting
from repositories.shift_repository import (
    get_entries_range,
    get_shift_entry_by_id,
    get_submission_entries_by_date,
)
from repositories.user_repository import get_all_users, move_active_user, set_user_active, update_user_name
from services.auth_service import (
    get_or_create_csrf_token,
    validate_csrf_or_400,
    get_client_ip,
    get_login_block_remaining,
    clear_login_failures,
    rotate_csrf_token,
    register_login_failure,
)
from services.confirmed_shift_service import (
    can_confirm_submission_entry,
    round_time_to_quarter,
    save_excluded_shift_from_entry,
    save_confirmed_shift_from_entry,
    validate_and_save_confirmed_shifts_bulk,
    validate_confirmed_shift_time,
    validate_submission_entry_for_exclusion,
    validate_submission_entry_for_confirmation,
)
from services.deadline_service import (
    get_active_deadline_config,
    get_deadline_settings_values,
    get_monthly_deadline_settings,
    get_submission_deadline,
    set_deadline_mode,
    set_submission_deadline,
    set_deadline_days_before,
    set_deadline_time,
    set_monthly_deadline_day,
    set_monthly_deadline_time,
    set_monthly_target,
)
from services.labor_summary_service import (
    DEFAULT_HOURLY_WAGE,
    build_daily_summary,
    build_daily_summary_from_rows,
    build_monthly_summary,
    get_labor_settings,
    get_hourly_wage,
    save_labor_settings,
    set_hourly_wage,
)
from services.summary_service import calculate_staff_summary, calculate_staff_summary_from_rows
from utils import (
    parse_ymd,
    to_ymd,
    parse_submission_deadline,
    parse_int_or_none,
    to_datetime_local_value,
    format_submission_deadline,
    format_relative_deadline,
    daterange_inclusive,
    html_escape,
    get_weekday_jp,
    is_valid_time_hhmm,
    today_jst,
)


admin_bp = Blueprint("admin", __name__)

DAILY_SHIFT_GRAPH_START_TIME_KEY = "DAILY_SHIFT_GRAPH_START_TIME"
DAILY_SHIFT_GRAPH_END_TIME_KEY = "DAILY_SHIFT_GRAPH_END_TIME"
DAILY_SHIFT_GRAPH_RANGE_KEY_PREFIX = "DAILY_SHIFT_GRAPH_RANGE_"
DEFAULT_DAILY_SHIFT_GRAPH_START_TIME = "08:00"
DEFAULT_DAILY_SHIFT_GRAPH_END_TIME = "23:00"


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    error_html = ""
    csrf_token_value = get_or_create_csrf_token()
    if request.method == "POST":
        csrf_error = validate_csrf_or_400()
        if csrf_error:
            return csrf_error
        ip_addr = get_client_ip()
        remaining = get_login_block_remaining(ip_addr)
        if remaining > 0:
            minutes_left = max(1, remaining // 60)
            error_html = f'<div class="alert alert-danger py-2 mb-3">ログイン試行回数が上限に達しました。あと約 {minutes_left} 分お待ちください。</div>'
        else:
            pw = request.form.get("password", "")
            if hmac.compare_digest(pw, ADMIN_PASSWORD):
                clear_login_failures(ip_addr)
                session.clear()
                session["logged_in"] = True
                session.permanent = False
                rotate_csrf_token()
                return redirect("/admin")
            register_login_failure(ip_addr)
            error_html = '<div class="alert alert-danger py-2 mb-3">パスワードが正しくありません。</div>'

    response = make_response(f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
  <title>管理者ログイン</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
  <div class="container py-5" style="max-width: 460px;">
    <div class="card shadow-sm border-0">
      <div class="card-body p-4">
        <h2 class="mb-1">管理者ログイン</h2>
        <div class="text-muted mb-4">管理画面に入るためのパスワードを入力してください。</div>
        {error_html}
        <form method="POST">
          <input type="hidden" name="csrf_token" value="{csrf_token_value}">
          <div class="mb-3">
            <label class="form-label" for="password">パスワード</label>
            <input id="password" type="password" class="form-control" name="password" placeholder="パスワードを入力" autocomplete="current-password">
          </div>
          <button type="submit" class="btn btn-dark w-100">ログイン</button>
        </form>
        <div class="small text-muted mt-3">管理者パスワードは `.env` の設定に従って判定されます。</div>
      </div>
    </div>
  </div>
</body>
</html>
""")
    response.headers["Content-Type"] = "text/html; charset=utf-8"
    return response

@admin_bp.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/login")


def parse_month_start(month_text: str):
    if not month_text:
        return None
    try:
        return datetime.strptime(f"{month_text}-01", "%Y-%m-%d").date()
    except Exception:
        return None


def _get_admin_date_state():
    today = today_jst()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    confirm_date = (request.args.get("confirm_date") or request.args.get("date") or "").strip()
    confirm_sort = (request.args.get("sort") or "display").strip()
    if confirm_sort not in {"display", "submitted"}:
        confirm_sort = "display"

    start_d = parse_ymd(start) or today
    end_d = parse_ymd(end) or (start_d + timedelta(days=6))
    if start_d > end_d:
        start_d, end_d = end_d, start_d

    start_value = to_ymd(start_d)
    end_value = to_ymd(end_d)
    confirm_date_value = to_ymd(parse_ymd(confirm_date) or start_d)
    current_month_start = date(today.year, today.month, 1)
    if today.month == 12:
        next_month_start = date(today.year + 1, 1, 1)
    else:
        next_month_start = date(today.year, today.month + 1, 1)
    if next_month_start.month == 12:
        month_after_next_start = date(next_month_start.year + 1, 1, 1)
    else:
        month_after_next_start = date(next_month_start.year, next_month_start.month + 1, 1)
    return {
        "today": today,
        "start_d": start_d,
        "end_d": end_d,
        "start_value": start_value,
        "end_value": end_value,
        "confirm_date_value": confirm_date_value,
        "confirm_sort": confirm_sort,
        "current_month_value": start_d.strftime("%Y-%m"),
        "current_month_start_value": to_ymd(current_month_start),
        "current_month_end_value": to_ymd(next_month_start - timedelta(days=1)),
        "next_month_start_value": to_ymd(next_month_start),
        "next_month_end_value": to_ymd(month_after_next_start - timedelta(days=1)),
    }


def _redirect_to_admin_page(default_path="/admin", include_confirm_date=False):
    next_page = (request.form.get("next_page") or "").strip()
    allowed_pages = {"/admin", "/admin/users", "/admin/deadline", "/admin/cost", "/admin/confirm", "/admin/daily-shift"}
    target_path = next_page if next_page in allowed_pages else default_path

    start = (request.form.get("start") or "").strip()
    end = (request.form.get("end") or "").strip()
    confirm_date = (request.form.get("confirm_date") or "").strip()
    confirm_sort = (request.form.get("sort") or "").strip()

    params = {}
    if start:
        params["start"] = start
    if end:
        params["end"] = end
    if include_confirm_date and confirm_date:
        params["confirm_date"] = confirm_date
    if confirm_sort:
        params["sort"] = confirm_sort

    query = urlencode(params)
    return redirect(f"{target_path}?{query}" if query else target_path)


def _build_admin_shell_context(active_nav: str):
    state = _get_admin_date_state()
    state["csrf_token_value"] = get_or_create_csrf_token()
    state["active_nav"] = active_nav
    return state


def _build_users_page_context():
    ctx = _build_admin_shell_context("users")
    admin_users = get_all_users(include_inactive=True)
    ctx["active_admin_users"] = [user for user in admin_users if int(user["active"]) == 1]
    ctx["inactive_admin_users"] = [user for user in admin_users if int(user["active"]) != 1]
    return ctx


def _build_deadline_page_context():
    ctx = _build_admin_shell_context("deadline")
    deadline_settings = get_deadline_settings_values()
    active_deadline_config = get_active_deadline_config(deadline_settings)
    submission_deadline = get_submission_deadline(deadline_settings)
    monthly_settings = get_monthly_deadline_settings(deadline_settings)
    mode_labels = {
        "fixed": "固定日時方式",
        "relative": "相対期限方式",
        "monthly_fixed": "月固定締切方式",
        "": "未設定",
    }
    ctx.update({
        "active_deadline_config": active_deadline_config,
        "submission_deadline_text": format_submission_deadline(submission_deadline),
        "submission_deadline_value": to_datetime_local_value(submission_deadline),
        "deadline_days_before_value": "" if active_deadline_config["relative_days_before"] is None else str(active_deadline_config["relative_days_before"]),
        "deadline_time_value": active_deadline_config["relative_time"] or "",
        "deadline_mode_value": active_deadline_config["mode"] or "",
        "monthly_deadline_day_value": "" if monthly_settings["day"] is None else str(monthly_settings["day"]),
        "monthly_target_value": monthly_settings["target"] or "next_month",
        "monthly_deadline_time_value": monthly_settings["time"] or "",
        "submission_status_text": mode_labels.get(active_deadline_config["mode"] or "", "未設定"),
        "relative_deadline_text": format_relative_deadline(
            active_deadline_config["relative_days_before"],
            active_deadline_config["relative_time"],
        ),
    })
    return ctx


def _build_cost_page_context():
    ctx = _build_admin_shell_context("cost")
    ctx["current_hourly_wage"] = get_hourly_wage()
    ctx["labor_settings"] = get_labor_settings()
    return ctx


def _build_home_page_context():
    ctx = _build_admin_shell_context("home")
    today = ctx["today"]
    today_value = to_ymd(today)

    confirmed_today = [
        row for row in get_confirmed_shifts_by_date(today_value)
        if int(row["is_assigned"] or 0) == 1 and row["start_time"] and row["end_time"]
    ]
    daily_labor = build_daily_summary_from_rows(today, confirmed_today)
    gross_minutes = sum(int(user["gross_minutes"] or 0) for user in daily_labor["users"])

    (
        ctx["daily_shift_graph_start_time"],
        ctx["daily_shift_graph_end_time"],
    ) = _auto_daily_shift_graph_range(confirmed_today)
    ctx["timeline_slots"] = _build_timeline_slots(ctx["daily_shift_graph_start_time"], ctx["daily_shift_graph_end_time"])
    ctx["today_timeline_shifts"] = [
        {
            "id": int(shift["id"]),
            "name": shift["name"] or "",
            "start_time": round_time_to_quarter(shift["start_time"] or ""),
            "end_time": round_time_to_quarter(shift["end_time"] or ""),
            "bar": _build_timeline_bar(
                round_time_to_quarter(shift["start_time"] or ""),
                round_time_to_quarter(shift["end_time"] or ""),
                ctx["timeline_slots"],
            ),
            "cells": _build_timeline_cells(
                round_time_to_quarter(shift["start_time"] or ""),
                round_time_to_quarter(shift["end_time"] or ""),
                ctx["timeline_slots"],
            ),
        }
        for shift in confirmed_today
    ]

    ctx.update({
        "today_value": today_value,
        "today_date_label": f"{today.year}年{today.month}月{today.day}日（{get_weekday_jp(today)}）",
        "today_working_count": len(confirmed_today),
        "today_labor_cost": daily_labor["total_labor_cost"],
        "today_total_work_hours": f"{gross_minutes / 60:.1f}h",
        "daily_shift_url": f"/admin/daily-shift?target_date={today_value}",
        "confirm_today_url": f"/admin/confirm?start={today_value}&end={today_value}&confirm_date={today_value}&sort=display",
    })
    return ctx


def _build_confirm_status_display(entry, decision=None, user=None):
    name = (entry["name"] if entry else user["name"]) or ""
    start_time = entry["start_time"] or "" if entry else ""
    end_time = entry["end_time"] or "" if entry else ""
    is_off = int(entry["off"]) == 1 if entry else False
    submitted = bool(entry)
    submitted_text = "\u672a\u63d0\u51fa"
    if entry:
        submitted_text = "\u4f11\u307f" if is_off else f"{start_time} - {end_time}" if start_time and end_time else "\u6642\u9593\u672a\u5165\u529b"

    confirmed_start_time = round_time_to_quarter(start_time)
    confirmed_end_time = round_time_to_quarter(end_time)
    confirmed_text = ""
    status_value = ""
    label = "\u672a\u51e6\u7406"
    badge_class = "secondary"
    bg_class = "bg-light"
    border_class = "border-light-subtle"
    is_dimmed = False

    if decision and int(decision["is_assigned"]) == 0:
        status_value = "rest"
        label = "\u4f11\u307f"
        badge_class = "secondary"
        bg_class = "bg-secondary-subtle"
        border_class = "border-secondary-subtle"
        confirmed_start_time = ""
        confirmed_end_time = ""
        confirmed_text = "\u4f11\u307f"
        is_dimmed = True
    elif decision and int(decision["is_assigned"]) == 1:
        confirmed_start_time = round_time_to_quarter(decision["start_time"] or "")
        confirmed_end_time = round_time_to_quarter(decision["end_time"] or "")
        confirmed_text = f"{confirmed_start_time} - {confirmed_end_time}"
        status_value = "assigned"
        label = "\u51fa\u52e4"
        badge_class = "success"
        bg_class = "bg-success-subtle"
        border_class = "border-success-subtle"
        is_dimmed = False

    return {
        "id": int(entry["id"]) if entry else None,
        "user_id": int(entry["user_id"]) if entry else int(user["id"]),
        "name": name,
        "off": is_off,
        "submitted": submitted,
        "start_time": start_time,
        "end_time": end_time,
        "submitted_text": submitted_text,
        "confirmed_start_time": confirmed_start_time,
        "confirmed_end_time": confirmed_end_time,
        "confirmed_text": confirmed_text,
        "confirmed_status_value": status_value,
        "confirmed_status_label": label,
        "confirmed_status_class": badge_class,
        "status_bg_class": bg_class,
        "status_border_class": border_class,
        "is_dimmed": is_dimmed,
        "can_confirm": True,
    }


def _sort_confirm_submission_entries(entries, active_users, submission_by_user_id, sort_mode):
    if sort_mode != "submitted":
        return entries
    display_index_by_user_id = {
        int(user["id"]): index
        for index, user in enumerate(active_users)
    }

    def sort_key(item):
        user_id = int(item["user_id"])
        entry = submission_by_user_id.get(user_id)
        display_index = display_index_by_user_id.get(user_id, 999999)
        submitted_at = (entry["updated_at"] or "") if entry else ""

        if item.get("confirmed_status_value") == "assigned":
            group = 0
        elif item.get("confirmed_status_value") == "rest":
            group = 1
        elif entry and int(entry["off"]) == 0 and entry["start_time"] and entry["end_time"]:
            group = 0
        elif entry and int(entry["off"]) == 1:
            group = 1
        else:
            group = 2

        return (group, submitted_at or "9999-12-31 23:59:59", display_index)

    return sorted(entries, key=sort_key)


def _build_period_status_counts(entries):
    submission_counts = {
        "work": 0,
        "rest": 0,
        "none": 0,
    }
    confirmation_counts = {
        "work": 0,
        "rest": 0,
        "pending": 0,
    }
    for item in entries:
        if item.get("submitted"):
            if item.get("off"):
                submission_counts["rest"] += 1
            elif item.get("start_time") and item.get("end_time"):
                submission_counts["work"] += 1
            else:
                submission_counts["none"] += 1
        else:
            submission_counts["none"] += 1

        if item.get("confirmed_status_value") == "assigned":
            confirmation_counts["work"] += 1
        elif item.get("confirmed_status_value") == "rest":
            confirmation_counts["rest"] += 1
        else:
            confirmation_counts["pending"] += 1

    return submission_counts, confirmation_counts

def get_daily_shift_graph_start_time(settings_values=None):
    raw_value = settings_values.get(DAILY_SHIFT_GRAPH_START_TIME_KEY) if settings_values is not None else get_setting(DAILY_SHIFT_GRAPH_START_TIME_KEY)
    value = (raw_value or "").strip()
    return value if is_valid_time_hhmm(value) else DEFAULT_DAILY_SHIFT_GRAPH_START_TIME


def get_daily_shift_graph_end_time(settings_values=None):
    raw_value = settings_values.get(DAILY_SHIFT_GRAPH_END_TIME_KEY) if settings_values is not None else get_setting(DAILY_SHIFT_GRAPH_END_TIME_KEY)
    value = (raw_value or "").strip()
    return value if is_valid_time_hhmm(value) else DEFAULT_DAILY_SHIFT_GRAPH_END_TIME


def _daily_shift_graph_range_key(date_value: str) -> str:
    return f"{DAILY_SHIFT_GRAPH_RANGE_KEY_PREFIX}{date_value}"


def _minutes_to_hhmm(minutes: int) -> str:
    normalized = int(minutes) % (24 * 60)
    return f"{normalized // 60:02d}:{normalized % 60:02d}"


def _get_saved_daily_shift_graph_range(date_value: str):
    raw_value = (get_setting(_daily_shift_graph_range_key(date_value)) or "").strip()
    if not raw_value or "|" not in raw_value:
        return None
    start_time, end_time = [part.strip() for part in raw_value.split("|", 1)]
    if not is_valid_time_hhmm(start_time) or not is_valid_time_hhmm(end_time):
        return None
    minutes_range = _time_range_minutes(start_time, end_time)
    if not minutes_range or minutes_range[1] - minutes_range[0] > 24 * 60:
        return None
    return start_time, end_time


def _auto_daily_shift_graph_range(confirmed_shifts):
    earliest_start = None
    latest_end = None
    for shift in confirmed_shifts:
        start_minutes = _time_to_minutes(shift["start_time"] or "")
        end_minutes = _time_to_minutes(shift["end_time"] or "")
        if start_minutes is None or end_minutes is None:
            continue
        if end_minutes <= start_minutes:
            end_minutes += 24 * 60
        earliest_start = start_minutes if earliest_start is None else min(earliest_start, start_minutes)
        latest_end = end_minutes if latest_end is None else max(latest_end, end_minutes)
    if earliest_start is None or latest_end is None:
        return get_daily_shift_graph_start_time(), get_daily_shift_graph_end_time()
    return _minutes_to_hhmm(earliest_start - 60), _minutes_to_hhmm(latest_end + 60)


def get_daily_shift_graph_range_for_date(date_value: str, confirmed_shifts):
    saved_range = _get_saved_daily_shift_graph_range(date_value)
    if saved_range:
        return {
            "start_time": saved_range[0],
            "end_time": saved_range[1],
            "is_saved": True,
        }
    start_time, end_time = _auto_daily_shift_graph_range(confirmed_shifts)
    return {
        "start_time": start_time,
        "end_time": end_time,
        "is_saved": False,
    }


def _time_range_minutes(start_time: str, end_time: str):
    start_minutes = _time_to_minutes(start_time)
    end_minutes = _time_to_minutes(end_time)
    if start_minutes is None or end_minutes is None:
        return None
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    return start_minutes, end_minutes


def _is_quarter_hour_value(time_text: str):
    return is_valid_time_hhmm(time_text or "") and int(time_text[3:]) in {0, 15, 30, 45}


def _normalize_time_for_range(time_text: str, range_start_minutes: int):
    minutes = _time_to_minutes(time_text)
    if minutes is None:
        return None
    if minutes < range_start_minutes:
        minutes += 24 * 60
    return minutes


def _build_timeline_slots(start_time: str, end_time: str):
    minutes_range = _time_range_minutes(start_time, end_time)
    if not minutes_range:
        minutes_range = _time_range_minutes(DEFAULT_DAILY_SHIFT_GRAPH_START_TIME, DEFAULT_DAILY_SHIFT_GRAPH_END_TIME)
    start_minutes, end_minutes = minutes_range
    slots = []
    current = start_minutes
    index = 0
    while current < end_minutes:
        label = _minutes_to_daily_sheet_label(current, start_minutes)
        slots.append({
            "label": label,
            "start": current,
            "end": min(current + 30, end_minutes),
            "hour_band": index // 2,
        })
        current += 30
        index += 1
    return slots


def _build_timeline_cells(start_time: str, end_time: str, slots):
    if not slots:
        return []
    range_start = int(slots[0]["start"])
    start_minutes = _normalize_time_for_range(start_time, range_start)
    end_minutes = _normalize_time_for_range(end_time, range_start)
    if start_minutes is None or end_minutes is None:
        return [{"active": False, "hour_band": slot["hour_band"]} for slot in slots]
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    return [
        {
            "active": slot["start"] < end_minutes and slot["end"] > start_minutes,
            "hour_band": slot["hour_band"],
        }
        for slot in slots
    ]


def _build_timeline_bar(start_time: str, end_time: str, slots):
    if not slots:
        return {"visible": False, "left": 0, "width": 0}
    range_start = int(slots[0]["start"])
    range_end = int(slots[-1]["end"])
    range_minutes = range_end - range_start
    if range_minutes <= 0:
        return {"visible": False, "left": 0, "width": 0}

    start_minutes = _normalize_time_for_range(start_time, range_start)
    end_minutes = _normalize_time_for_range(end_time, range_start)
    if start_minutes is None or end_minutes is None:
        return {"visible": False, "left": 0, "width": 0}
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60

    visible_start = max(start_minutes, range_start)
    visible_end = min(end_minutes, range_end)
    if visible_end <= visible_start:
        return {"visible": False, "left": 0, "width": 0}

    return {
        "visible": True,
        "left": round(((visible_start - range_start) / range_minutes) * 100, 4),
        "width": round(((visible_end - visible_start) / range_minutes) * 100, 4),
    }


def _build_confirm_page_context():
    ctx = _build_admin_shell_context("confirm")
    ctx["quarter_hour_options"] = [
        f"{hour:02d}:{minute:02d}"
        for hour in range(24)
        for minute in (0, 15, 30, 45)
    ]
    timeline_settings = get_settings([DAILY_SHIFT_GRAPH_START_TIME_KEY, DAILY_SHIFT_GRAPH_END_TIME_KEY])
    ctx["daily_shift_graph_start_time"] = get_daily_shift_graph_start_time(timeline_settings)
    ctx["daily_shift_graph_end_time"] = get_daily_shift_graph_end_time(timeline_settings)
    ctx["timeline_slots"] = _build_timeline_slots(ctx["daily_shift_graph_start_time"], ctx["daily_shift_graph_end_time"])
    daily_shift_date = parse_ymd(request.args.get("daily_shift_date") or session.get("daily_shift_date") or "")
    ctx["daily_shift_date_value"] = to_ymd(daily_shift_date) if daily_shift_date else ctx["confirm_date_value"]
    rows = get_entries_range(ctx["start_value"], ctx["end_value"])
    active_users = [user for user in get_all_users(include_inactive=False) if int(user["active"]) == 1]
    summary_by_date = calculate_staff_summary_from_rows(ctx["start_d"], ctx["end_d"], rows, len(active_users))
    submission_entries = [
        row for row in rows
        if row["date"] == ctx["confirm_date_value"] and int(row["active"]) == 1
    ]
    decision_rows = get_confirmed_shift_decisions_range(ctx["start_value"], ctx["end_value"])
    if not (ctx["start_value"] <= ctx["confirm_date_value"] <= ctx["end_value"]):
        decision_rows = list(decision_rows) + list(get_confirmed_shift_decisions_by_date(ctx["confirm_date_value"]))
    confirmed_shifts = get_confirmed_shifts_by_date(ctx["confirm_date_value"])

    by_date = {}
    for row in rows:
        if int(row["active"]) != 1:
            continue
        by_date.setdefault(row["date"], []).append(row)

    decision_map = {}
    decision_user_ids_by_date = {}
    for row in decision_rows:
        key = (int(row["user_id"]), row["date"])
        decision_map[key] = row
        decision_user_ids_by_date.setdefault(row["date"], set()).add(int(row["user_id"]))

    period_days = []
    for d in daterange_inclusive(ctx["start_d"], ctx["end_d"]):
        ymd = to_ymd(d)
        summary = summary_by_date[ymd]
        entries = []
        active_user_ids = {int(user["id"]) for user in active_users}
        day_entries_by_user_id = {
            int(row["user_id"]): row
            for row in by_date.get(ymd, [])
        }
        for user in active_users:
            user_id = int(user["id"])
            row = day_entries_by_user_id.get(user_id)
            decision = decision_map.get((user_id, ymd))
            entries.append(_build_confirm_status_display(row, decision, user))

        if summary["required"] <= 0:
            status_label = "必要人数未設定"
            status_class = "secondary"
        elif summary["is_shortage"]:
            status_label = f"不足 {summary['shortage_count']}人"
            status_class = "danger"
        else:
            status_label = "充足"
            status_class = "success"

        submission_counts, confirmation_counts = _build_period_status_counts(entries)

        if not active_user_ids:
            confirm_status_label = "提出なし"
            confirm_status_class = "secondary"
        elif confirmation_counts["pending"] == 0:
            confirm_status_label = "全員確定済み"
            confirm_status_class = "success"
        else:
            confirm_status_label = "未処理あり"
            confirm_status_class = "warning"

        period_days.append({
            "ymd": ymd,
            "date_label": f"{d.month}/{d.day}（{get_weekday_jp(d)}）",
            "required": summary["required"],
            "status_label": status_label,
            "status_class": status_class,
            "working_count": summary["working_count"],
            "off_count": summary["off_count"],
            "not_submitted_count": summary["not_submitted_count"],
            "submission_work_count": submission_counts["work"],
            "submission_rest_count": submission_counts["rest"],
            "submission_none_count": submission_counts["none"],
            "confirmed_work_count": confirmation_counts["work"],
            "confirmed_rest_count": confirmation_counts["rest"],
            "confirmed_pending_count": confirmation_counts["pending"],
            "confirm_status_label": confirm_status_label,
            "confirm_status_class": confirm_status_class,
            "entries": entries,
        })

    submission_by_user_id = {int(entry["user_id"]): entry for entry in submission_entries}
    ctx["submission_entries"] = []
    for user in active_users:
        entry = submission_by_user_id.get(int(user["id"]))
        decision = decision_map.get((int(user["id"]), ctx["confirm_date_value"]))
        ctx["submission_entries"].append(_build_confirm_status_display(entry, decision, user))
    ctx["submission_entries"] = _sort_confirm_submission_entries(
        ctx["submission_entries"],
        active_users,
        submission_by_user_id,
        ctx["confirm_sort"],
    )

    ctx["editable_submission_entries"] = []
    for entry in submission_entries:
        decision = decision_map.get((int(entry["user_id"]), entry["date"]))
        ctx["editable_submission_entries"].append(_build_confirm_status_display(entry, decision))
    ctx["active_users"] = [
        {"id": int(user["id"]), "name": user["name"] or ""}
        for user in active_users
    ]
    ctx["confirmed_shifts"] = [
        {
            "id": int(shift["id"]),
            "user_id": int(shift["user_id"]),
            "name": shift["name"] or "",
            "start_time": round_time_to_quarter(shift["start_time"] or ""),
            "end_time": round_time_to_quarter(shift["end_time"] or ""),
            "cells": _build_timeline_cells(
                round_time_to_quarter(shift["start_time"] or ""),
                round_time_to_quarter(shift["end_time"] or ""),
                ctx["timeline_slots"],
            ),
        }
        for shift in confirmed_shifts
    ]
    ctx["period_days"] = period_days
    return ctx


def _get_daily_shift_target_date():
    today = today_jst()
    raw_value = (request.args.get("target_date") or request.args.get("confirm_date") or session.get("daily_shift_date") or "").strip()
    target_date = parse_ymd(raw_value) or today
    session["daily_shift_date"] = to_ymd(target_date)
    return target_date


def _build_daily_shift_page_context():
    target_date = _get_daily_shift_target_date()
    target_value = to_ymd(target_date)
    ctx = _build_admin_shell_context("daily_shift")
    ctx["target_date_value"] = target_value
    ctx["quarter_hour_options"] = [
        f"{hour:02d}:{minute:02d}"
        for hour in range(24)
        for minute in (0, 15, 30, 45)
    ]
    active_users = [user for user in get_all_users(include_inactive=False) if int(user["active"]) == 1]
    confirmed_shifts = get_confirmed_shifts_by_date(target_value)
    graph_range = get_daily_shift_graph_range_for_date(target_value, confirmed_shifts)
    ctx["daily_shift_graph_start_time"] = graph_range["start_time"]
    ctx["daily_shift_graph_end_time"] = graph_range["end_time"]
    ctx["daily_shift_graph_range_is_saved"] = graph_range["is_saved"]
    ctx["timeline_slots"] = _build_timeline_slots(ctx["daily_shift_graph_start_time"], ctx["daily_shift_graph_end_time"])
    confirmed_user_ids = {int(shift["user_id"]) for shift in confirmed_shifts}
    ctx["active_users"] = [
        {
            "id": int(user["id"]),
            "name": user["name"] or "",
            "already_confirmed": int(user["id"]) in confirmed_user_ids,
        }
        for user in active_users
    ]
    ctx["confirmed_shifts"] = [
        {
            "id": int(shift["id"]),
            "user_id": int(shift["user_id"]),
            "name": shift["name"] or "",
            "start_time": round_time_to_quarter(shift["start_time"] or ""),
            "end_time": round_time_to_quarter(shift["end_time"] or ""),
            "cells": _build_timeline_cells(
                round_time_to_quarter(shift["start_time"] or ""),
                round_time_to_quarter(shift["end_time"] or ""),
                ctx["timeline_slots"],
            ),
        }
        for shift in confirmed_shifts
        if shift["start_time"] and shift["end_time"]
    ]
    ctx["confirm_page_url"] = f"/admin/confirm?start={target_value}&end={target_value}&confirm_date={target_value}"
    return ctx


def _extract_html_fragment(full_html: str, tag_name: str) -> str:
    match = re.search(rf"<{tag_name}[^>]*>(.*)</{tag_name}>", full_html, flags=re.DOTALL | re.IGNORECASE)
    return match.group(1).strip() if match else ""


@admin_bp.route("/admin/users", methods=["GET"])
def admin_users_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("admin_users.html", **_build_users_page_context()))


@admin_bp.route("/admin/deadline", methods=["GET"])
def admin_deadline_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("admin_deadline.html", **_build_deadline_page_context()))


@admin_bp.route("/admin/cost", methods=["GET"])
def admin_cost_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("admin_cost.html", **_build_cost_page_context()))


@admin_bp.route("/admin/confirm", methods=["GET"])
def admin_confirm_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("admin_confirm.html", **_build_confirm_page_context()))


@admin_bp.route("/admin/daily-shift", methods=["GET"])
def admin_daily_shift_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("admin_daily_shift.html", **_build_daily_shift_page_context()))


@admin_bp.route("/admin/manual", methods=["GET"])
def admin_manual_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("manual/index.html", **_build_admin_shell_context("manual")))


@admin_bp.route("/admin/manual/pdf", methods=["GET"])
def admin_manual_pdf_page():
    if not session.get("logged_in"):
        return redirect("/login")
    return make_response(render_template("manual/pdf.html"))


@admin_bp.route("/admin/confirm_shift", methods=["POST"])
def admin_confirm_shift():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    entry_id = parse_int_or_none(request.form.get("entry_id"))
    confirmed_start_time = (request.form.get("confirmed_start_time") or "").strip()
    confirmed_end_time = (request.form.get("confirmed_end_time") or "").strip()
    if not entry_id:
        return "entry_id が不正です", 400

    entry = get_shift_entry_by_id(entry_id)
    error_message = validate_submission_entry_for_confirmation(entry, confirmed_start_time, confirmed_end_time)
    if error_message:
        return error_message, 400

    save_confirmed_shift_from_entry(entry, confirmed_start_time, confirmed_end_time)

    return _redirect_to_admin_page(include_confirm_date=True)


@admin_bp.route("/admin/exclude_shift", methods=["POST"])
def admin_exclude_shift():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    entry_id = parse_int_or_none(request.form.get("entry_id"))
    if not entry_id:
        return "entry_id が不正です", 400

    entry = get_shift_entry_by_id(entry_id)
    error_message = validate_submission_entry_for_exclusion(entry)
    if error_message:
        return error_message, 400

    save_excluded_shift_from_entry(entry)
    return _redirect_to_admin_page(include_confirm_date=True)


@admin_bp.route("/admin/save_confirmed_shifts_bulk", methods=["POST"])
def admin_save_confirmed_shifts_bulk():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    confirm_date = (request.form.get("confirm_date") or "").strip()
    if not parse_ymd(confirm_date):
        return "confirm_date が不正です", 400

    submitted_items = []
    for raw_user_id in request.form.getlist("user_id"):
        user_id = parse_int_or_none(raw_user_id)
        if not user_id:
            return "user_id が不正です", 400
        entry_id = parse_int_or_none(request.form.get(f"entry_id_{user_id}"))
        submitted_items.append({
            "user_id": user_id,
            "entry_id": entry_id,
            "status": (request.form.get(f"status_{user_id}") or "").strip(),
            "start_time": request.form.get(f"start_time_{user_id}") or "",
            "end_time": request.form.get(f"end_time_{user_id}") or "",
        })

    save_unsubmitted_as_rest = request.form.get("save_unsubmitted_as_rest") == "1"
    error_message, status_code = validate_and_save_confirmed_shifts_bulk(
        confirm_date,
        submitted_items,
        save_unsubmitted_as_rest=save_unsubmitted_as_rest,
    )
    if error_message:
        return error_message, status_code

    return _redirect_to_admin_page(default_path="/admin/confirm", include_confirm_date=True)


@admin_bp.route("/admin/save_daily_shift_sheet_settings", methods=["POST"])
@admin_bp.route("/admin/daily-shift/settings", methods=["POST"])
def admin_save_daily_shift_sheet_settings():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    start_time = (request.form.get("daily_shift_graph_start_time") or "").strip()
    end_time = (request.form.get("daily_shift_graph_end_time") or "").strip()
    daily_shift_date = (request.form.get("daily_shift_date") or request.form.get("target_date") or "").strip()
    reset_to_auto = (request.form.get("action") or "").strip() == "reset_auto"
    if daily_shift_date and not parse_ymd(daily_shift_date):
        return "1日シフト表の対象日が不正です", 400
    if reset_to_auto:
        if daily_shift_date:
            delete_setting(_daily_shift_graph_range_key(daily_shift_date))
            session["daily_shift_date"] = daily_shift_date
        next_page = (request.form.get("next_page") or "").strip()
        target_path = next_page if next_page in {"/admin", "/admin/confirm", "/admin/daily-shift"} else "/admin/confirm"
        params = {}
        for key in ("start", "end", "confirm_date"):
            value = (request.form.get(key) or "").strip()
            if value:
                params[key] = value
        if daily_shift_date:
            params["daily_shift_date"] = daily_shift_date
            if target_path == "/admin/daily-shift":
                params["target_date"] = daily_shift_date
        query = urlencode(params)
        return redirect(f"{target_path}?{query}" if query else target_path)
    if not is_valid_time_hhmm(start_time) or not is_valid_time_hhmm(end_time):
        return "表示開始時間 / 表示終了時間が不正です", 400

    if daily_shift_date and not parse_ymd(daily_shift_date):
        return "1日シフト表の対象日が不正です", 400

    minutes_range = _time_range_minutes(start_time, end_time)
    if not minutes_range:
        return "表示開始時間 / 表示終了時間が不正です", 400
    start_minutes, end_minutes = minutes_range
    if end_minutes - start_minutes > 24 * 60:
        return "表示範囲は最大24時間以内にしてください", 400

    if daily_shift_date:
        upsert_setting(_daily_shift_graph_range_key(daily_shift_date), f"{start_time}|{end_time}")
        session["daily_shift_date"] = daily_shift_date
    else:
        upsert_setting(DAILY_SHIFT_GRAPH_START_TIME_KEY, start_time)
        upsert_setting(DAILY_SHIFT_GRAPH_END_TIME_KEY, end_time)

    next_page = (request.form.get("next_page") or "").strip()
    target_path = next_page if next_page in {"/admin", "/admin/confirm", "/admin/daily-shift"} else "/admin/confirm"
    params = {}
    for key in ("start", "end", "confirm_date"):
        value = (request.form.get(key) or "").strip()
        if value:
            params[key] = value
    if daily_shift_date:
        params["daily_shift_date"] = daily_shift_date
        if target_path == "/admin/daily-shift":
            params["target_date"] = daily_shift_date
    query = urlencode(params)
    return redirect(f"{target_path}?{query}" if query else target_path)


@admin_bp.route("/admin/save_timeline_shifts", methods=["POST"])
@admin_bp.route("/admin/daily-shift/save", methods=["POST"])
def admin_save_timeline_shifts():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    confirm_date = (request.form.get("confirm_date") or "").strip()
    if not parse_ymd(confirm_date):
        return "confirm_date が不正です", 400

    graph_start_time = (request.form.get("daily_shift_graph_start_time") or "").strip()
    graph_end_time = (request.form.get("daily_shift_graph_end_time") or "").strip()
    if not is_valid_time_hhmm(graph_start_time) or not is_valid_time_hhmm(graph_end_time):
        return "表示開始時間 / 表示終了時間が不正です", 400
    graph_range = _time_range_minutes(graph_start_time, graph_end_time)
    if not graph_range:
        return "表示開始時間 / 表示終了時間が不正です", 400
    if graph_range[1] - graph_range[0] > 24 * 60:
        return "表示範囲は最大24時間以内にしてください", 400

    active_user_ids = {int(user["id"]) for user in get_all_users(include_inactive=False) if int(user["active"]) == 1}
    saved_user_ids = set()
    for row_id in request.form.getlist("timeline_row_id"):
        confirmed_shift_id = parse_int_or_none(request.form.get(f"timeline_confirmed_shift_id_{row_id}"))
        user_id = parse_int_or_none(request.form.get(f"timeline_user_id_{row_id}"))
        start_time = (request.form.get(f"timeline_start_time_{row_id}") or "").strip()
        end_time = (request.form.get(f"timeline_end_time_{row_id}") or "").strip()
        delete_flag = request.form.get(f"timeline_delete_{row_id}") == "1"

        if confirmed_shift_id and delete_flag:
            existing = get_confirmed_shift_by_id(confirmed_shift_id)
            if not existing or existing["date"] != confirm_date:
                return "確定シフトが見つかりません", 404
            upsert_confirmed_shift(
                user_id=int(existing["user_id"]),
                date_str=confirm_date,
                start_time="",
                end_time="",
                source_entry_id=existing["source_entry_id"],
                is_assigned=0,
            )
            continue
        if not user_id and not start_time and not end_time:
            continue
        if not user_id or user_id not in active_user_ids:
            return "スタッフが不正です", 400
        if user_id in saved_user_ids:
            return "同じスタッフが複数行に入っています", 400
        saved_user_ids.add(user_id)
        if not (is_valid_time_hhmm(start_time) and is_valid_time_hhmm(end_time)):
            return "勤務開始 / 勤務終了が不正です", 400
        if not (_is_quarter_hour_value(start_time) and _is_quarter_hour_value(end_time)):
            return "勤務時間は15分単位で選択してください", 400
        if start_time == end_time:
            return "勤務開始と勤務終了は異なる時刻を選択してください", 400

        source_entry_id = None
        if confirmed_shift_id:
            existing = get_confirmed_shift_by_id(confirmed_shift_id)
            if not existing or existing["date"] != confirm_date:
                return "確定シフトが見つかりません", 404
            source_entry_id = existing["source_entry_id"]
        upsert_confirmed_shift(
            user_id=user_id,
            date_str=confirm_date,
            start_time=start_time,
            end_time=end_time,
            source_entry_id=source_entry_id,
            is_assigned=1,
        )

    return _redirect_to_admin_page(default_path="/admin/confirm", include_confirm_date=True)


@admin_bp.route("/admin/update_confirmed_shift", methods=["POST"])
def admin_update_confirmed_shift():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    confirmed_shift_id = parse_int_or_none(request.form.get("confirmed_shift_id"))
    confirmed_start_time = (request.form.get("confirmed_start_time") or "").strip()
    confirmed_end_time = (request.form.get("confirmed_end_time") or "").strip()
    if not confirmed_shift_id:
        return "confirmed_shift_id が不正です", 400

    confirmed_shift = get_confirmed_shift_by_id(confirmed_shift_id)
    if not confirmed_shift:
        return "確定シフトが見つかりません", 404

    error_message = validate_confirmed_shift_time(confirmed_start_time, confirmed_end_time)
    if error_message:
        return error_message, 400

    upsert_confirmed_shift(
        user_id=int(confirmed_shift["user_id"]),
        date_str=confirmed_shift["date"],
        start_time=confirmed_start_time,
        end_time=confirmed_end_time,
        source_entry_id=confirmed_shift["source_entry_id"],
    )

    return _redirect_to_admin_page(include_confirm_date=True)


@admin_bp.route("/admin/delete_confirmed_shift", methods=["POST"])
def admin_delete_confirmed_shift():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    confirmed_shift_id = parse_int_or_none(request.form.get("confirmed_shift_id"))
    if not confirmed_shift_id:
        return "confirmed_shift_id が不正です", 400

    confirmed_shift = get_confirmed_shift_by_id(confirmed_shift_id)
    if not confirmed_shift:
        return "確定シフトが見つかりません", 404

    delete_confirmed_shift(confirmed_shift_id)

    return _redirect_to_admin_page(include_confirm_date=True)


@admin_bp.route("/admin/update_hourly_wage", methods=["POST"])
def admin_update_hourly_wage():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    raw_wage = (request.form.get("base_hourly_wage") or request.form.get("hourly_wage") or "").strip()
    hourly_wage = parse_int_or_none(raw_wage)
    if hourly_wage is None or hourly_wage < 0:
        return "時給の形式が不正です", 400

    if "base_hourly_wage" in request.form:
        save_labor_settings(
            hourly_wage,
            request.form.get("break_enabled") == "on",
            request.form.get("overtime_enabled") == "on",
            request.form.get("night_enabled") == "on",
        )
    else:
        set_hourly_wage(hourly_wage)

    return _redirect_to_admin_page(include_confirm_date=True)


@admin_bp.route("/admin/api/daily_summary", methods=["GET"])
def admin_api_daily_summary():
    if not session.get("logged_in"):
        return jsonify({"error": "ログインが必要です"}), 401

    date_text = (request.args.get("date") or "").strip()
    target_date = parse_ymd(date_text)
    if not target_date:
        return jsonify({"error": "date の形式が不正です"}), 400

    return jsonify(build_daily_summary(target_date))


@admin_bp.route("/admin/api/monthly_summary", methods=["GET"])
def admin_api_monthly_summary():
    if not session.get("logged_in"):
        return jsonify({"error": "ログインが必要です"}), 401

    month_text = (request.args.get("month") or "").strip()
    start_date = parse_month_start(month_text)
    if not start_date:
        return jsonify({"error": "month の形式が不正です"}), 400

    if start_date.month == 12:
        end_exclusive_date = date(start_date.year + 1, 1, 1)
    else:
        end_exclusive_date = date(start_date.year, start_date.month + 1, 1)

    return jsonify(build_monthly_summary(start_date, end_exclusive_date, month_text))

@admin_bp.route("/admin/update_required", methods=["POST"])
def admin_update_required():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    start = request.form.get("start", "").strip()
    end = request.form.get("end", "").strip()

    start_d = parse_ymd(start)
    end_d = parse_ymd(end)
    if not start_d or not end_d:
        return "日付が不正です", 400

    if start_d > end_d:
        start_d, end_d = end_d, start_d

    for d in daterange_inclusive(start_d, end_d):
        ymd = to_ymd(d)
        raw = request.form.get(f"required_{ymd}", "").strip()
        if raw == "":
            required_count = 0
        else:
            try:
                required_count = int(raw)
            except Exception:
                required_count = 0

        if required_count < 0:
            required_count = 0

        upsert_required_staff(ymd, required_count)

    return _redirect_to_admin_page(include_confirm_date=True)

@admin_bp.route("/admin/update_submission_deadline", methods=["POST"])
def admin_update_submission_deadline():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    start = request.form.get("start", "").strip()
    end = request.form.get("end", "").strip()
    deadline_mode = (request.form.get("deadline_mode") or "").strip()
    deadline_raw = request.form.get("submission_deadline", "").strip()
    days_before_raw = request.form.get("deadline_days_before", "").strip()
    deadline_time = (request.form.get("deadline_time") or "").strip()
    monthly_day_raw = (request.form.get("monthly_deadline_day") or "").strip()
    monthly_target = (request.form.get("monthly_target") or "next_month").strip()
    monthly_deadline_time = (request.form.get("monthly_deadline_time") or "").strip()
    deadline = parse_submission_deadline(deadline_raw)
    days_before = parse_int_or_none(days_before_raw) if days_before_raw != "" else None
    monthly_day = parse_int_or_none(monthly_day_raw) if monthly_day_raw != "" else None

    if deadline_mode not in ("", "fixed", "relative", "monthly_fixed"):
        return "提出期限の設定方式が不正です", 400
    if deadline_raw and not deadline:
        return "提出期限の形式が不正です", 400
    if days_before_raw != "" and days_before is None:
        return "何日前の形式が不正です", 400
    if days_before is not None and days_before < 0:
        return "何日前は0以上で入力してください", 400
    if deadline_time and not is_valid_time_hhmm(deadline_time):
        return "締切時刻の形式が不正です（HH:MM）", 400
    if deadline_mode == "monthly_fixed" and monthly_day_raw == "":
        return "締切日を入力してください", 400
    if monthly_day_raw != "" and monthly_day is None:
        return "締切日は1〜31の範囲で入力してください", 400
    if deadline_mode == "monthly_fixed" and monthly_day is not None and not (1 <= monthly_day <= 31):
        return "締切日は1～31の範囲で入力してください", 400
    if monthly_day is not None and not (1 <= monthly_day <= 31):
        return "月固定締切日は1〜31で入力してください", 400
    if monthly_target not in ("current_month", "next_month"):
        return "月固定締切の対象が不正です", 400
    if monthly_deadline_time and not is_valid_time_hhmm(monthly_deadline_time):
        return "月固定締切時刻の形式が不正です（HH:MM）", 400
    if deadline_mode == "fixed" and deadline is None:
        return "固定日時方式を使う場合は締切日時を入力してください", 400
    if deadline_mode == "relative" and (days_before is None or not deadline_time):
        return "相対期限方式を使う場合は何日前と締切時刻を入力してください", 400
    if deadline_mode == "monthly_fixed" and (monthly_day is None or not monthly_deadline_time):
        return "月固定締切方式を使う場合は締切日と締切時刻を入力してください", 400

    set_deadline_mode(deadline_mode or "none")
    set_submission_deadline(deadline)
    set_deadline_days_before(days_before)
    set_deadline_time(deadline_time or None)
    set_monthly_deadline_day(monthly_day)
    set_monthly_target(monthly_target)
    set_monthly_deadline_time(monthly_deadline_time or None)

    return _redirect_to_admin_page()


@admin_bp.route("/admin/update_deadline_mode", methods=["POST"])
def admin_update_deadline_mode():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    deadline_mode = (request.form.get("deadline_mode") or "").strip()
    if deadline_mode not in ("", "fixed", "relative", "monthly_fixed"):
        return "提出期限の設定方式が不正です", 400

    set_deadline_mode(deadline_mode or "none")
    return _redirect_to_admin_page(default_path="/admin/deadline")


@admin_bp.route("/admin/users/deactivate", methods=["POST"])
def admin_users_deactivate():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    user_id = parse_int_or_none(request.form.get("user_id"))
    if not user_id:
        return "user_id が不正です", 400

    updated = set_user_active(user_id, 0)
    if updated <= 0:
        return "対象ユーザーが見つかりません", 404
    return _redirect_to_admin_page()

@admin_bp.route("/admin/users/activate", methods=["POST"])
def admin_users_activate():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    user_id = parse_int_or_none(request.form.get("user_id"))
    if not user_id:
        return "user_id が不正です", 400

    updated = set_user_active(user_id, 1)
    if updated <= 0:
        return "対象ユーザーが見つかりません", 404
    return _redirect_to_admin_page()

@admin_bp.route("/admin/users/update_name", methods=["POST"])
def admin_users_update_name():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error

    user_id = parse_int_or_none(request.form.get("user_id"))
    new_name = (request.form.get("new_name") or "").strip()
    if not user_id:
        return "user_id が不正です", 400
    if not new_name:
        return "new_name は必須です", 400

    updated = update_user_name(user_id, new_name)
    if updated <= 0:
        return "対象ユーザーが見つかりません", 404

    return _redirect_to_admin_page()


@admin_bp.route("/admin/users/<int:user_id>/move_up", methods=["POST"])
def admin_users_move_up(user_id):
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error
    move_active_user(user_id, "up")
    return _redirect_to_admin_page(default_path="/admin/users")


@admin_bp.route("/admin/users/<int:user_id>/move_down", methods=["POST"])
def admin_users_move_down(user_id):
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_error = validate_csrf_or_400()
    if csrf_error:
        return csrf_error
    move_active_user(user_id, "down")
    return _redirect_to_admin_page(default_path="/admin/users")

@admin_bp.route("/admin/home_content", methods=["GET"])
def admin_home_content():
    if not session.get("logged_in"):
        return redirect("/login")
    csrf_token_value = get_or_create_csrf_token()

    today = today_jst()
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()

    start_d = parse_ymd(start) or today
    end_d = parse_ymd(end) or (start_d + timedelta(days=6))
    if start_d > end_d:
        start_d, end_d = end_d, start_d

    confirm_date = request.args.get("confirm_date", "").strip()
    start_value = to_ymd(start_d)
    end_value = to_ymd(end_d)
    confirm_date_value = to_ymd(parse_ymd(confirm_date) or start_d)
    active_deadline_config = get_active_deadline_config()
    submission_deadline = get_submission_deadline()
    submission_deadline_text = format_submission_deadline(submission_deadline)
    submission_deadline_value = to_datetime_local_value(submission_deadline)
    deadline_days_before_value = (
        "" if active_deadline_config["relative_days_before"] is None else str(active_deadline_config["relative_days_before"])
    )
    deadline_time_value = active_deadline_config["relative_time"] or ""
    deadline_mode_value = active_deadline_config["mode"] or ""
    current_hourly_wage = get_hourly_wage()
    current_month_value = start_d.strftime("%Y-%m")
    deadline_mode_labels = {
        "fixed": "固定日時方式",
        "relative": "相対期限方式",
        "": "未設定",
    }
    submission_status_text = deadline_mode_labels.get(deadline_mode_value, "未設定")
    submission_status_badge = (
        '<span class="badge bg-warning text-dark">未設定</span>'
        if not active_deadline_config["is_configured"]
        else f'<span class="badge bg-primary">{html_escape(submission_status_text)}</span>'
    )

    rows = get_entries_range(start_value, end_value)
    summary_by_date = calculate_staff_summary(start_d, end_d)
    submission_entries = get_submission_entries_by_date(confirm_date_value)
    confirmed_shifts = get_confirmed_shifts_by_date(confirm_date_value)
    admin_users = get_all_users(include_inactive=True)
    active_admin_users = [user for user in admin_users if int(user["active"]) == 1]
    inactive_admin_users = [user for user in admin_users if int(user["active"]) != 1]

    by_date = {}
    for r in rows:
        if int(r["active"]) != 1:
            continue
        by_date.setdefault(r["date"], []).append(r)

    table_rows = ""
    for d in daterange_inclusive(start_d, end_d):
        ymd = to_ymd(d)
        items = by_date.get(ymd, [])
        summary = summary_by_date[ymd]

        if summary["required"] <= 0:
            status_badge = '<span class="badge bg-secondary">必要人数未設定</span>'
        elif summary["is_shortage"]:
            status_badge = f'<span class="badge bg-danger">不足 {summary["shortage_count"]}人</span>'
        else:
            status_badge = '<span class="badge bg-success">充足</span>'

        if not items:
            submit_html = "<div class='text-muted'>（提出なし）</div>"
        else:
            lines = []
            for r in items:
                name = html_escape(r["name"] or "")
                if int(r["off"]) == 1:
                    lines.append(
                        f"<div class='border rounded p-2 mb-2 bg-white'>"
                        f"<b>{name}</b>：<span class='badge bg-secondary'>休み</span>"
                        f"</div>"
                    )
                else:
                    st = html_escape(r["start_time"] or "")
                    et = html_escape(r["end_time"] or "")
                    lines.append(
                        f"<div class='border rounded p-2 mb-2 bg-white'>"
                        f"<b>{name}</b>：<span class='badge bg-primary'>{st}-{et}</span>"
                        f"</div>"
                    )
            submit_html = "".join(lines)

        table_rows += f"""
        <tr>
          <td style="width:220px;">
            <div class="fw-semibold">{d.month}/{d.day}（{get_weekday_jp(d)}）</div>
            <div class="small text-muted">{ymd}</div>
          </td>
          <td style="width:180px;">
            <input type="number"
                   class="form-control form-control-sm"
                   min="0"
                   name="required_{ymd}"
                   value="{summary["required"]}">
          </td>
          <td style="width:220px;">
            <div class="mb-1">{status_badge}</div>
            <div class="small">
              <div>出勤予定：<b>{summary["working_count"]}</b>人</div>
              <div>休み：<b>{summary["off_count"]}</b>人</div>
              <div>未提出：<b>{summary["not_submitted_count"]}</b>人</div>
            </div>
          </td>
          <td>{submit_html}</td>
        </tr>
        """

    active_user_table_rows = ""
    for user in active_admin_users:
        active_user_table_rows += f"""
        <tr>
          <td>{html_escape(user["name"] or "")}</td>
          <td><code>{html_escape(user["line_user_id"] or "")}</code></td>
          <td><span class="badge bg-success">有効</span></td>
          <td>
            <div class="d-grid gap-2">
              <form method="POST" action="/admin/users/deactivate" onsubmit="return confirm('このユーザーを無効化しますか？');">
                <input type="hidden" name="csrf_token" value="{csrf_token_value}">
                <input type="hidden" name="start" value="{start_value}">
                <input type="hidden" name="end" value="{end_value}">
                <input type="hidden" name="user_id" value="{user["id"]}">
                <button type="submit" class="btn btn-sm btn-outline-danger">無効化</button>
              </form>
              <form method="POST" action="/admin/users/update_name" class="d-flex gap-2">
                <input type="hidden" name="csrf_token" value="{csrf_token_value}">
                <input type="hidden" name="start" value="{start_value}">
                <input type="hidden" name="end" value="{end_value}">
                <input type="hidden" name="user_id" value="{user["id"]}">
                <input type="text" class="form-control form-control-sm" name="new_name" value="{html_escape(user["name"] or "")}" placeholder="新しい名前">
                <button type="submit" class="btn btn-sm btn-outline-primary">更新</button>
              </form>
            </div>
          </td>
        </tr>
        """
    if not active_user_table_rows:
        active_user_table_rows = """
        <tr>
          <td colspan="4" class="text-muted">有効ユーザーはまだいません。</td>
        </tr>
        """

    inactive_user_table_rows = ""
    for user in inactive_admin_users:
        inactive_user_table_rows += f"""
        <tr>
          <td>{html_escape(user["name"] or "")}</td>
          <td><code>{html_escape(user["line_user_id"] or "")}</code></td>
          <td><span class="badge bg-secondary">無効</span></td>
          <td>
            <div class="d-grid gap-2">
              <form method="POST" action="/admin/users/activate" onsubmit="return confirm('このユーザーを再有効化しますか？');">
                <input type="hidden" name="csrf_token" value="{csrf_token_value}">
                <input type="hidden" name="start" value="{start_value}">
                <input type="hidden" name="end" value="{end_value}">
                <input type="hidden" name="user_id" value="{user["id"]}">
                <button type="submit" class="btn btn-sm btn-outline-success">再有効化</button>
              </form>
              <form method="POST" action="/admin/users/update_name" class="d-flex gap-2">
                <input type="hidden" name="csrf_token" value="{csrf_token_value}">
                <input type="hidden" name="start" value="{start_value}">
                <input type="hidden" name="end" value="{end_value}">
                <input type="hidden" name="user_id" value="{user["id"]}">
                <input type="text" class="form-control form-control-sm" name="new_name" value="{html_escape(user["name"] or "")}" placeholder="新しい名前">
                <button type="submit" class="btn btn-sm btn-outline-primary">更新</button>
              </form>
            </div>
          </td>
        </tr>
        """
    if not inactive_user_table_rows:
        inactive_user_table_rows = """
        <tr>
          <td colspan="4" class="text-muted">無効ユーザーはいません。</td>
        </tr>
        """

    submission_shift_rows = ""
    for entry in submission_entries:
        name = html_escape(entry["name"] or "")
        if int(entry["off"]) == 1:
            submission_label = '<span class="badge bg-secondary">休み</span>'
            action_html = '<div class="text-muted small">休みの提出です</div>'
        elif not entry["start_time"] or not entry["end_time"]:
            submission_label = '<span class="badge bg-warning text-dark">時間未入力</span>'
            action_html = '<div class="text-muted small">開始/終了時間が未入力のため確定できません</div>'
        else:
            start_time = html_escape(entry["start_time"])
            end_time = html_escape(entry["end_time"])
            submission_label = f'<span class="badge bg-primary">{start_time}-{end_time}</span>'
            action_html = f'<a class="btn btn-sm btn-outline-primary" href="/admin/confirm?start={start_value}&end={end_value}&confirm_date={confirm_date_value}">確定シフト管理で設定</a>'

        submission_shift_rows += f"""
        <tr>
          <td>{name}</td>
          <td>{submission_label}</td>
          <td>{action_html}</td>
        </tr>
        """
    if not submission_shift_rows:
        submission_shift_rows = """
        <tr>
          <td colspan="3" class="text-muted">この日の提出シフトはありません。</td>
        </tr>
        """

    confirmed_shift_rows = ""
    for confirmed_shift in confirmed_shifts:
        confirmed_shift_rows += f"""
        <tr>
          <td>{html_escape(confirmed_shift["name"] or "")}</td>
          <td>
            <form method="POST" action="/admin/update_confirmed_shift" class="row g-2">
              <input type="hidden" name="csrf_token" value="{csrf_token_value}">
              <input type="hidden" name="start" value="{start_value}">
              <input type="hidden" name="end" value="{end_value}">
              <input type="hidden" name="confirm_date" value="{confirm_date_value}">
              <input type="hidden" name="confirmed_shift_id" value="{confirmed_shift["id"]}">
              <div class="col-md-4">
                <input type="time" class="form-control form-control-sm" name="confirmed_start_time" step="900" value="{html_escape(confirmed_shift["start_time"] or "")}">
              </div>
              <div class="col-md-4">
                <input type="time" class="form-control form-control-sm" name="confirmed_end_time" step="900" value="{html_escape(confirmed_shift["end_time"] or "")}">
              </div>
              <div class="col-md-4 d-grid">
                <button type="submit" class="btn btn-sm btn-outline-primary">更新</button>
              </div>
            </form>
          </td>
          <td>
            <form method="POST" action="/admin/delete_confirmed_shift" onsubmit="return confirm('この確定シフトを削除しますか？');">
              <input type="hidden" name="csrf_token" value="{csrf_token_value}">
              <input type="hidden" name="start" value="{start_value}">
              <input type="hidden" name="end" value="{end_value}">
              <input type="hidden" name="confirm_date" value="{confirm_date_value}">
              <input type="hidden" name="confirmed_shift_id" value="{confirmed_shift["id"]}">
              <button type="submit" class="btn btn-sm btn-outline-danger">削除</button>
            </form>
          </td>
        </tr>
        """
    if not confirmed_shift_rows:
        confirmed_shift_rows = """
        <tr>
          <td colspan="3" class="text-muted">この日の確定シフトはありません。</td>
        </tr>
        """

    return f"""<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Shift Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
  body {{
    background: #f5f7fb;
  }}
  .hero-card {{
    border: 0;
    box-shadow: 0 10px 30px rgba(0,0,0,.06);
    border-radius: 18px;
  }}
  .section-card {{
    border: 0;
    box-shadow: 0 8px 24px rgba(0,0,0,.05);
    border-radius: 18px;
  }}
  .summary-box {{
    border-radius: 16px;
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
    border: 1px solid rgba(0,0,0,.06);
    padding: 16px;
    height: 100%;
  }}
  .table thead th {{
    white-space: nowrap;
  }}
</style>
</head>
<body>
<div class="container py-4 py-md-5">

  <div class="card hero-card mb-4">
    <div class="card-body p-4 p-md-5">
      <div class="d-flex flex-column flex-md-row justify-content-between align-items-md-center gap-3">
        <div>
          <div class="text-uppercase small text-muted mb-2">Shift Management Dashboard</div>
          <h2 class="mb-1">シフト管理画面</h2>
          <div class="text-muted">提出状況、必要人数、不足状況をまとめて確認できます。</div>
        </div>
        <div class="d-flex gap-2">
          <a class="btn btn-success" href="/admin_export?start={start_value}&end={end_value}">Excel出力</a>
          <a class="btn btn-outline-secondary" href="/logout">ログアウト</a>
        </div>
      </div>
    </div>
  </div>

  <div class="row g-3 mb-4">
    <div class="col-md-4">
      <div class="summary-box">
        <div class="text-muted small">表示期間</div>
        <div class="fs-5 fw-semibold">{start_value} 〜 {end_value}</div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="summary-box">
        <div class="text-muted small">対象日数</div>
        <div class="fs-5 fw-semibold">{(end_d - start_d).days + 1}日間</div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="summary-box">
        <div class="text-muted small">機能</div>
        <div class="fs-6 fw-semibold">必要人数設定 / 不足確認 / Excel出力</div>
      </div>
    </div>
  </div>

  <div class="card section-card mb-4">
    <div class="card-body">
      <div class="d-flex flex-column flex-md-row justify-content-between align-items-md-center gap-3 mb-3">
        <div>
          <h5 class="card-title mb-1">登録ユーザー一覧</h5>
          <div class="text-muted small">退職者などは無効化にすると、未提出人数などの運用対象から外れます。過去のシフト履歴は削除されません。</div>
        </div>
      </div>

      <div class="table-responsive">
        <table class="table table-striped align-middle">
          <thead>
            <tr>
              <th>名前</th>
              <th>line_user_id</th>
              <th>状態</th>
              <th>操作</th>
            </tr>
          </thead>
          <tbody>
            {active_user_table_rows}
          </tbody>
        </table>
      </div>

      <details class="mt-3">
        <summary class="fw-semibold">無効ユーザーを表示</summary>
        <div class="table-responsive mt-3">
          <table class="table table-striped align-middle mb-0">
            <thead>
              <tr>
                <th>名前</th>
                <th>line_user_id</th>
                <th>状態</th>
                <th>操作</th>
              </tr>
            </thead>
            <tbody>
              {inactive_user_table_rows}
            </tbody>
          </table>
        </div>
      </details>
    </div>
  </div>

  <div class="card section-card mb-4">
    <div class="card-body">
      <div class="d-flex flex-column flex-md-row justify-content-between align-items-md-center gap-3 mb-3">
        <div>
          <h5 class="card-title mb-1">シフト提出期限</h5>
          <div class="text-muted small">固定日時方式と、シフト日の何日前かで締める相対期限方式を切り替えて設定できます。</div>
        </div>
        <div>{submission_status_badge}</div>
      </div>

      <div class="row g-3 mb-3">
        <div class="col-md-6">
          <div class="summary-box">
            <div class="text-muted small">現在有効な方式</div>
            <div class="fs-5 fw-semibold">{html_escape(submission_status_text)}</div>
            <div class="small text-muted mt-2">現在のルール: {html_escape(active_deadline_config["display"])}</div>
          </div>
        </div>
        <div class="col-md-6">
          <div class="summary-box">
            <div class="text-muted small">固定日時設定</div>
            <div class="fs-6 fw-semibold">{html_escape(submission_deadline_text)}</div>
            <div class="small text-muted mt-2">相対期限設定: {html_escape(format_relative_deadline(active_deadline_config["relative_days_before"], active_deadline_config["relative_time"]))}</div>
          </div>
        </div>
      </div>

      <form method="POST" action="/admin/update_submission_deadline" class="row g-3">
        <input type="hidden" name="csrf_token" value="{csrf_token_value}">
        <input type="hidden" name="start" value="{start_value}">
        <input type="hidden" name="end" value="{end_value}">
        <div class="col-12">
          <label class="form-label d-block mb-2">締切方式</label>
          <div class="d-flex flex-column flex-md-row gap-3">
            <div class="form-check">
              <input class="form-check-input" type="radio" name="deadline_mode" id="deadlineModeNone" value="" {"checked" if deadline_mode_value == "" else ""}>
              <label class="form-check-label" for="deadlineModeNone">未設定</label>
            </div>
            <div class="form-check">
              <input class="form-check-input" type="radio" name="deadline_mode" id="deadlineModeFixed" value="fixed" {"checked" if deadline_mode_value == "fixed" else ""}>
              <label class="form-check-label" for="deadlineModeFixed">固定日時方式</label>
            </div>
            <div class="form-check">
              <input class="form-check-input" type="radio" name="deadline_mode" id="deadlineModeRelative" value="relative" {"checked" if deadline_mode_value == "relative" else ""}>
              <label class="form-check-label" for="deadlineModeRelative">相対期限方式</label>
            </div>
          </div>
        </div>
        <div class="col-md-4">
          <label class="form-label">固定日時の締切</label>
          <input type="datetime-local" class="form-control" name="submission_deadline" value="{submission_deadline_value}">
          <div class="form-text">固定日時方式で使う締切です。</div>
        </div>
        <div class="col-md-3">
          <label class="form-label">何日前</label>
          <input type="number" class="form-control" name="deadline_days_before" min="0" value="{deadline_days_before_value}">
          <div class="form-text">相対期限方式で使います。</div>
        </div>
        <div class="col-md-2">
          <label class="form-label">締切時刻</label>
          <input type="time" class="form-control" name="deadline_time" value="{deadline_time_value}">
          <div class="form-text">例: 23:59</div>
        </div>
        <div class="col-md-3 d-flex align-items-end">
          <button type="submit" class="btn btn-primary w-100">提出期限を保存</button>
        </div>
        <div class="col-12">
          <div class="form-text">
            固定日時方式は全日共通の締切です。相対期限方式は「各シフト日のN日前 HH:MM」を締切として自動判定します。
          </div>
        </div>
      </form>
    </div>
  </div>

  <div class="card section-card mb-4">
    <div class="card-body">
      <div class="d-flex flex-column flex-md-row justify-content-between align-items-md-center gap-3 mb-3">
        <div>
          <h5 class="card-title mb-1">人件費集計</h5>
          <div class="text-muted small">確定シフトをもとに、日別・月別の概算人件費を確認できます。</div>
        </div>
      </div>

      <form method="POST" action="/admin/update_hourly_wage" class="row g-3 mb-4">
        <input type="hidden" name="csrf_token" value="{csrf_token_value}">
        <input type="hidden" name="start" value="{start_value}">
        <input type="hidden" name="end" value="{end_value}">
        <div class="col-md-4">
          <label class="form-label">時給設定</label>
          <input type="number" class="form-control" name="hourly_wage" min="0" value="{current_hourly_wage}">
          <div class="form-text">人件費集計で使う基本時給です。未設定時は {DEFAULT_HOURLY_WAGE} 円で計算します。</div>
        </div>
        <div class="col-md-3 d-flex align-items-end">
          <button type="submit" class="btn btn-outline-primary w-100">人件費設定を保存</button>
        </div>
      </form>

      <div class="row g-4">
        <div class="col-lg-6">
          <div class="border rounded-4 p-3 h-100 bg-light-subtle">
            <h6 class="mb-3">日別集計</h6>
            <div class="row g-2">
              <div class="col-sm-8">
                <input type="date" class="form-control" id="dailySummaryDate" value="{start_value}">
              </div>
              <div class="col-sm-4">
                <button type="button" class="btn btn-dark w-100" id="loadDailySummaryBtn">日別集計</button>
              </div>
            </div>
            <div id="dailySummaryResult" class="mt-3 small text-muted">日付を選んで集計してください。</div>
          </div>
        </div>
        <div class="col-lg-6">
          <div class="border rounded-4 p-3 h-100 bg-light-subtle">
            <h6 class="mb-3">月別集計</h6>
            <div class="row g-2">
              <div class="col-sm-8">
                <input type="month" class="form-control" id="monthlySummaryMonth" value="{current_month_value}">
              </div>
              <div class="col-sm-4">
                <button type="button" class="btn btn-dark w-100" id="loadMonthlySummaryBtn">月別集計</button>
              </div>
            </div>
            <div id="monthlySummaryResult" class="mt-3 small text-muted">年月を選んで集計してください。</div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <div class="card section-card mb-4">
    <div class="card-body">
      <div class="d-flex flex-column flex-md-row justify-content-between align-items-md-center gap-3 mb-3">
        <div>
          <h5 class="card-title mb-1">確定シフト管理</h5>
          <div class="text-muted small">提出された希望シフトと、管理者が調整した確定シフトを日付ごとに確認できます。</div>
        </div>
      </div>

      <form method="GET" class="row g-3 mb-4" id="homeConfirmDateForm">
        <input type="hidden" name="start" value="{start_value}">
        <input type="hidden" name="end" value="{end_value}">
        <div class="col-md-2 d-flex align-items-end">
          <button type="button" class="btn btn-outline-dark w-100" onclick="shiftTargetDate('homeConfirmDateInput', 'homeConfirmDateForm', -1)">前の日</button>
        </div>
        <div class="col-md-4">
          <label class="form-label">確定シフト対象日</label>
          <input type="date" class="form-control" name="confirm_date" id="homeConfirmDateInput" value="{confirm_date_value}">
        </div>
        <div class="col-md-2 d-flex align-items-end">
          <button type="button" class="btn btn-outline-dark w-100" onclick="shiftTargetDate('homeConfirmDateInput', 'homeConfirmDateForm', 1)">次の日</button>
        </div>
        <div class="col-md-3 d-flex align-items-end">
          <button type="submit" class="btn btn-outline-dark w-100">確定シフトを表示</button>
        </div>
        <div class="col-md-4 d-flex align-items-end">
          <a class="btn btn-success w-100" href="/admin/export_confirmed?start={start_value}&end={end_value}">確定シフトをExcel出力</a>
        </div>
      </form>

      <div class="row g-4">
        <div class="col-lg-6">
          <div class="border rounded-4 p-3 h-100 bg-light-subtle">
            <h6 class="mb-3">{confirm_date_value} 提出シフト</h6>
            <div class="table-responsive">
              <table class="table table-sm align-middle mb-0">
                <thead>
                  <tr>
                    <th>名前</th>
                    <th>希望内容</th>
                    <th>操作</th>
                  </tr>
                </thead>
                <tbody>
                  {submission_shift_rows}
                </tbody>
              </table>
            </div>
          </div>
        </div>
        <div class="col-lg-6">
          <div class="border rounded-4 p-3 h-100 bg-light-subtle">
            <h6 class="mb-3">{confirm_date_value} 確定シフト</h6>
            <div class="table-responsive">
              <table class="table table-sm align-middle mb-0">
                <thead>
                  <tr>
                    <th>名前</th>
                    <th>時間編集</th>
                    <th>削除</th>
                  </tr>
                </thead>
                <tbody>
                  {confirmed_shift_rows}
                </tbody>
              </table>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <div class="card section-card mb-4">
    <div class="card-body">
      <h5 class="card-title mb-3">表示期間</h5>
      <form method="GET" class="row g-3">
        <div class="col-md-4">
          <label class="form-label">開始日</label>
          <input type="date" class="form-control" name="start" value="{start_value}">
        </div>
        <div class="col-md-4">
          <label class="form-label">終了日</label>
          <input type="date" class="form-control" name="end" value="{end_value}">
        </div>
        <div class="col-md-2 d-flex align-items-end">
          <button type="submit" class="btn btn-dark w-100">表示</button>
        </div>
        <div class="col-md-2 d-flex align-items-end">
          <a class="btn btn-outline-dark w-100" href="/admin">今日から7日</a>
        </div>
      </form>
    </div>
  </div>

  <form method="POST" action="/admin/update_required">
    <input type="hidden" name="csrf_token" value="{csrf_token_value}">
    <input type="hidden" name="start" value="{start_value}">
    <input type="hidden" name="end" value="{end_value}">

    <div class="card section-card">
      <div class="card-body">
        <div class="d-flex flex-column flex-md-row justify-content-between align-items-md-center gap-3 mb-3">
          <div>
            <h5 class="card-title mb-1">提出一覧・必要人数設定</h5>
            <div class="text-muted small">必要人数を日ごとに入力して保存できます。必要人数に対して出勤予定人数が足りない日は自動で不足表示になります。</div>
          </div>
          <button type="submit" class="btn btn-primary">必要人数を保存</button>
        </div>

        <div class="table-responsive">
          <table class="table table-striped align-middle">
            <thead>
              <tr>
                <th>日付</th>
                <th>必要人数</th>
                <th>状況</th>
                <th>提出一覧</th>
              </tr>
            </thead>
            <tbody>
              {table_rows}
            </tbody>
          </table>
        </div>

        <div class="mt-3">
          <button type="submit" class="btn btn-primary">必要人数を保存</button>
        </div>
      </div>
    </div>
  </form>

</div>
<script>
function formatHours(value) {{
  const num = Number(value || 0);
  return num.toFixed(2);
}}

function formatYen(value) {{
  return new Intl.NumberFormat("ja-JP").format(Number(value || 0));
}}

function renderSummaryResult(data, labelKey) {{
  const users = Array.isArray(data.users) ? data.users : [];
  const detailBlocks = users.length
    ? users.map((user) => `
        <details class="mt-2">
          <summary class="fw-semibold">内訳を見る: ${{user.name || "未設定"}}</summary>
          <div class="mt-2 small">
            <div>勤務時間: ${{formatHours(user.hours)}} 時間</div>
            <div>勤務分数: ${{user.minutes}} 分</div>
            <div>給料: ¥${{formatYen(user.wage)}}</div>
          </div>
        </details>
      `).join("")
    : '<div class="text-muted small mt-2">対象データはありません。</div>';

  return `
    <div class="border rounded-3 bg-white p-3">
      <div class="fw-semibold mb-2">${{data[labelKey]}}</div>
      <div>時給: ¥${{formatYen(data.hourly_wage)}}</div>
      <div>総労働時間: ${{formatHours(data.total_hours)}} 時間</div>
      <div>総労働分数: ${{data.total_minutes}} 分</div>
      <div>総人件費: ¥${{formatYen(data.total_labor_cost)}}</div>
      <div class="mt-3">${{detailBlocks}}</div>
    </div>
  `;
}}

async function loadAdminSummary(url, resultId, labelKey) {{
  const resultEl = document.getElementById(resultId);
  resultEl.innerHTML = '<div class="text-muted">読み込み中...</div>';
  try {{
    const response = await fetch(url, {{
      headers: {{
        "X-Requested-With": "XMLHttpRequest"
      }}
    }});
    const data = await response.json().catch(() => ({{}}));
    if (!response.ok) {{
      throw new Error(data.error || "集計の取得に失敗しました");
    }}
    resultEl.innerHTML = renderSummaryResult(data, labelKey);
  }} catch (error) {{
    resultEl.innerHTML = `<div class="text-danger">${{error.message || error}}</div>`;
  }}
}}

document.getElementById("loadDailySummaryBtn").addEventListener("click", async () => {{
  const targetDate = document.getElementById("dailySummaryDate").value;
  await loadAdminSummary(`/admin/api/daily_summary?date=${{encodeURIComponent(targetDate)}}`, "dailySummaryResult", "date");
}});

document.getElementById("loadMonthlySummaryBtn").addEventListener("click", async () => {{
  const targetMonth = document.getElementById("monthlySummaryMonth").value;
  await loadAdminSummary(`/admin/api/monthly_summary?month=${{encodeURIComponent(targetMonth)}}`, "monthlySummaryResult", "month");
}});
</script>
</body>
</html>
"""

@admin_bp.route("/admin", methods=["GET"])
def admin():
    if not session.get("logged_in"):
        return redirect("/login")

    return make_response(render_template("admin_home.html", **_build_home_page_context()))


@admin_bp.route("/export", methods=["GET"])
@admin_bp.route("/admin_export", methods=["GET"])
def admin_export():
    if not session.get("logged_in"):
        return redirect("/login")

    start_d = parse_ymd(request.args.get("start"))
    end_d = parse_ymd(request.args.get("end"))
    if not start_d or not end_d:
        return "start / end が不正です", 400
    if start_d > end_d:
        return "start は end 以下で指定してください", 400

    rows = get_entries_range(to_ymd(start_d), to_ymd(end_d))
    required_map = get_required_staff_range(to_ymd(start_d), to_ymd(end_d))
    export_days = list(daterange_inclusive(start_d, end_d))

    user_names = [
        user["name"]
        for user in get_all_users(include_inactive=False)
        if int(user["active"]) == 1
    ]

    entry_map = {}
    for r in rows:
        if int(r["active"]) != 1:
            continue
        key = (r["name"], r["date"])
        if int(r["off"]) == 1:
            entry_map[key] = "休"
        elif r["start_time"] and r["end_time"]:
            entry_map[key] = f'{r["start_time"]}-{r["end_time"]}'
        else:
            entry_map[key] = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "シフト表"

    thin_side = Side(style="thin", color="B8C2CF")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    saturday_fill = PatternFill("solid", fgColor="DDEEFE")
    sunday_fill = PatternFill("solid", fgColor="F7E4DE")
    name_fill = PatternFill("solid", fgColor="F8FAFC")

    ws["A1"] = f"シフト表（{to_ymd(start_d)}〜{to_ymd(end_d)}）"
    ws["B1"] = ""
    ws["A1"].font = title_font
    ws["B1"].font = title_font

    ws["A4"] = "スタッフ"
    ws["A5"] = ""
    for row_idx in (4, 5):
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = center
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    weekday_labels = ["月", "火", "水", "木", "金", "土", "日"]

    for idx, current_date in enumerate(export_days, start=1):
        col = idx + 1
        weekday = current_date.weekday()
        fill = header_fill
        if weekday == 5:
            fill = saturday_fill
        elif weekday == 6:
            fill = sunday_fill

        ws.cell(row=4, column=col, value=current_date.day)
        ws.cell(row=5, column=col, value=weekday_labels[weekday])

        for row_idx in (4, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = center
            cell.font = header_font
            cell.fill = fill
            cell.border = border

    start_row = 6
    for idx, name in enumerate(user_names):
        row_num = start_row + idx
        name_cell = ws.cell(row=row_num, column=1, value=name)
        name_cell.alignment = left
        name_cell.font = Font(bold=True)
        name_cell.fill = name_fill
        name_cell.border = border

        for idx, current_date in enumerate(export_days, start=1):
            ymd = to_ymd(current_date)
            value = entry_map.get((name, ymd), "")
            cell = ws.cell(row=row_num, column=idx + 1, value=value)
            cell.alignment = cell_alignment
            cell.border = border

            weekday = current_date.weekday()
            if weekday == 5:
                cell.fill = saturday_fill
            elif weekday == 6:
                cell.fill = sunday_fill

    ws.freeze_panes = "B6"
    ws.column_dimensions["A"].width = 18
    for idx, _ in enumerate(export_days, start=1):
        ws.column_dimensions[get_column_letter(idx + 1)].width = 11

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 22
    for row_num in range(start_row, start_row + max(len(user_names), 1)):
        ws.row_dimensions[row_num].height = 28

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"shift_{to_ymd(start_d)}_{to_ymd(end_d)}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/admin/export_confirmed", methods=["GET"])
def admin_export_confirmed():
    if not session.get("logged_in"):
        return redirect("/login")

    start_d = parse_ymd(request.args.get("start"))
    end_d = parse_ymd(request.args.get("end"))
    if not start_d or not end_d:
        return "start / end \u304c\u4e0d\u6b63\u3067\u3059", 400
    if start_d > end_d:
        return "start \u306f end \u4ee5\u4e0b\u3067\u6307\u5b9a\u3057\u3066\u304f\u3060\u3055\u3044", 400

    rows = get_confirmed_shifts_range(to_ymd(start_d), to_ymd(end_d))
    export_days = list(daterange_inclusive(start_d, end_d))
    by_date = {}
    for row in rows:
        by_date.setdefault(row["date"], []).append(row)
    for ymd, day_rows in by_date.items():
        by_date[ymd] = sorted(day_rows, key=_confirmed_export_sort_key)

    daily_costs = {}
    for current_date in export_days:
        daily_costs[to_ymd(current_date)] = build_daily_summary(current_date)["total_labor_cost"]

    wb = Workbook()
    ws = wb.active
    ws.title = "\u78ba\u5b9a\u30b7\u30d5\u30c8"

    thin_side = Side(style="thin", color="B8C2CF")
    medium_side = Side(style="medium", color="6B7280")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    section_border = Border(left=medium_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    subheader_font = Font(bold=True)
    date_fill = PatternFill("solid", fgColor="374151")
    subheader_fill = PatternFill("solid", fgColor="E8EEF7")
    saturday_fill = PatternFill("solid", fgColor="DDEEFE")
    sunday_fill = PatternFill("solid", fgColor="F7E4DE")
    empty_fill = PatternFill("solid", fgColor="F8FAFC")
    cost_fill = PatternFill("solid", fgColor="EAF7EA")

    total_columns = 7 * 3
    ws.cell(row=1, column=1, value=f"\u78ba\u5b9a\u30b7\u30d5\u30c8\u4e00\u89a7\uff08{to_ymd(start_d)} \uff5e {to_ymd(end_d)}\uff09")
    ws.cell(row=1, column=1).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws.cell(row=2, column=1, value="\u8868\u793a\u671f\u9593\u306e\u78ba\u5b9a\u6e08\u307f\u30b7\u30d5\u30c8\u306e\u307f\u3092\u30011\u9031\u9593\u3054\u3068\u306b\u6a2a\u4e26\u3073\u3067\u8868\u793a\u3057\u3066\u3044\u307e\u3059\u3002")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)

    row_cursor = 4
    week_index = 1
    for week_start in range(0, len(export_days), 7):
        week_days = export_days[week_start:week_start + 7]
        week_height = max(1, max(len(by_date.get(to_ymd(day), [])) for day in week_days))

        ws.cell(row=row_cursor, column=1, value=f"\u7b2c{week_index}\u9031")
        ws.cell(row=row_cursor, column=1).font = subheader_font
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=total_columns)
        row_cursor += 1

        for day_index, current_date in enumerate(week_days):
            ymd = to_ymd(current_date)
            start_col = day_index * 3 + 1
            end_col = start_col + 2
            weekday = current_date.weekday()
            date_cell = ws.cell(
                row=row_cursor,
                column=start_col,
                value=f"{current_date.month}/{current_date.day}\uff08{get_weekday_jp(current_date)}\uff09\n{ymd}",
            )
            date_cell.font = header_font
            date_cell.alignment = center
            date_cell.fill = date_fill
            for col in range(start_col, end_col + 1):
                ws.cell(row=row_cursor, column=col).fill = date_fill
                ws.cell(row=row_cursor, column=col).border = section_border if col == start_col else border
            ws.merge_cells(start_row=row_cursor, start_column=start_col, end_row=row_cursor, end_column=end_col)

            day_fill = saturday_fill if weekday == 5 else sunday_fill if weekday == 6 else subheader_fill
            for offset, label in enumerate(("\u540d\u524d", "\u958b\u59cb", "\u7d42\u4e86")):
                cell = ws.cell(row=row_cursor + 1, column=start_col + offset, value=label)
                cell.font = subheader_font
                cell.alignment = center
                cell.fill = day_fill
                cell.border = section_border if offset == 0 else border

            day_rows = by_date.get(ymd, [])
            if day_rows:
                for item_index in range(week_height):
                    target_row = row_cursor + 2 + item_index
                    if item_index < len(day_rows):
                        item = day_rows[item_index]
                        values = (item["name"] or "", item["start_time"] or "", item["end_time"] or "")
                    else:
                        values = ("", "", "")
                    for offset, value in enumerate(values):
                        cell = ws.cell(row=target_row, column=start_col + offset, value=value)
                        cell.alignment = left if offset == 0 else center
                        cell.border = section_border if offset == 0 else border
            else:
                no_shift_row = row_cursor + 2
                ws.cell(row=no_shift_row, column=start_col, value="\u78ba\u5b9a\u30b7\u30d5\u30c8\u306a\u3057")
                ws.cell(row=no_shift_row, column=start_col).alignment = center
                ws.cell(row=no_shift_row, column=start_col).fill = empty_fill
                for col in range(start_col, end_col + 1):
                    ws.cell(row=no_shift_row, column=col).border = section_border if col == start_col else border
                    ws.cell(row=no_shift_row, column=col).fill = empty_fill
                ws.merge_cells(start_row=no_shift_row, start_column=start_col, end_row=no_shift_row, end_column=end_col)
                for extra_index in range(1, week_height):
                    target_row = row_cursor + 2 + extra_index
                    for col in range(start_col, end_col + 1):
                        ws.cell(row=target_row, column=col).border = section_border if col == start_col else border

            cost_row = row_cursor + 2 + week_height
            cost_cell = ws.cell(row=cost_row, column=start_col, value=f"\u6982\u7b97\u4eba\u4ef6\u8cbb: \u00a5{daily_costs.get(ymd, 0):,}")
            cost_cell.font = subheader_font
            cost_cell.alignment = center
            for col in range(start_col, end_col + 1):
                ws.cell(row=cost_row, column=col).fill = cost_fill
                ws.cell(row=cost_row, column=col).border = section_border if col == start_col else border
            ws.merge_cells(start_row=cost_row, start_column=start_col, end_row=cost_row, end_column=end_col)

        row_cursor += week_height + 4
        week_index += 1

    for day_index in range(7):
        start_col = day_index * 3 + 1
        ws.column_dimensions[get_column_letter(start_col)].width = 18
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 9
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 9
    for row_num in range(1, row_cursor + 1):
        ws.row_dimensions[row_num].height = 24
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"confirmed_shifts_{to_ymd(start_d)}_{to_ymd(end_d)}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def _confirmed_export_sort_key(row):
    start_minutes = _time_to_minutes(row["start_time"] or "")
    end_minutes = _time_to_minutes(row["end_time"] or "")
    display_order = row["display_order"] if "display_order" in row.keys() else None
    try:
        display_order_value = int(display_order) if display_order is not None else 999999
    except (TypeError, ValueError):
        display_order_value = 999999

    return (
        start_minutes if start_minutes is not None else 999999,
        end_minutes if end_minutes is not None else 999999,
        display_order_value,
        row["name"] or "",
    )

def _time_to_minutes(time_text: str):
    if not time_text:
        return None
    try:
        hour, minute = [int(part) for part in str(time_text).split(":", 1)]
    except (TypeError, ValueError):
        return None
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return hour * 60 + minute


def _minutes_to_daily_sheet_label(minutes: int, graph_start_minutes: int) -> str:
    day_minutes = minutes % (24 * 60)
    label = f"{day_minutes // 60:02d}:{day_minutes % 60:02d}"
    return f"翌{label}" if minutes >= 24 * 60 else label


def _daily_sheet_time_label(time_text: str, graph_start_time: str, start_time_text: str = "") -> str:
    minutes = _time_to_minutes(time_text)
    if minutes is None:
        return ""
    graph_start_minutes = _time_to_minutes(graph_start_time) or 0
    start_minutes = _time_to_minutes(start_time_text)
    is_next_day = False
    if start_minutes is not None and minutes <= start_minutes:
        is_next_day = True
    if is_next_day:
        minutes += 24 * 60
    return _minutes_to_daily_sheet_label(minutes, graph_start_minutes)


def _build_daily_sheet_time_options(start_time: str, end_time: str, step_minutes: int):
    minutes_range = _time_range_minutes(start_time, end_time)
    if not minutes_range:
        minutes_range = _time_range_minutes(DEFAULT_DAILY_SHIFT_GRAPH_START_TIME, DEFAULT_DAILY_SHIFT_GRAPH_END_TIME)
    start_minutes, end_minutes = minutes_range
    labels = []
    current = start_minutes
    while current <= end_minutes:
        labels.append(_minutes_to_daily_sheet_label(current, start_minutes))
        current += step_minutes
    if labels[-1] != _minutes_to_daily_sheet_label(end_minutes, start_minutes):
        labels.append(_minutes_to_daily_sheet_label(end_minutes, start_minutes))
    return labels


def _daily_sheet_label_to_numeric_time(time_label: str):
    if not time_label:
        return None
    is_next_day = time_label.startswith("翌")
    raw = time_label[1:] if is_next_day else time_label
    minutes = _time_to_minutes(raw)
    if minutes is None:
        return None
    return minutes / (24 * 60)


def _normalize_minutes_for_daily_sheet(time_text: str, range_start_minutes: int):
    minutes = _time_to_minutes(time_text)
    if minutes is None:
        return None
    if minutes < range_start_minutes:
        minutes += 24 * 60
    return minutes


def _clip_daily_shift_bar(start_time: str, end_time: str, range_start_minutes: int, range_end_minutes: int):
    start_minutes = _time_to_minutes(start_time)
    end_minutes = _time_to_minutes(end_time)
    if start_minutes is None or end_minutes is None:
        return None
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60

    candidates = (
        (start_minutes, end_minutes),
        (start_minutes + 24 * 60, end_minutes + 24 * 60),
    )
    best_range = None
    best_overlap = 0
    for candidate_start, candidate_end in candidates:
        overlap = min(candidate_end, range_end_minutes) - max(candidate_start, range_start_minutes)
        if overlap > best_overlap:
            best_overlap = overlap
            best_range = (candidate_start, candidate_end)
    if not best_range:
        return None

    clipped_start = max(best_range[0], range_start_minutes)
    clipped_end = min(best_range[1], range_end_minutes)
    if clipped_end <= clipped_start:
        return None
    return clipped_start, clipped_end


@admin_bp.route("/admin/export_daily_shift_sheet", methods=["GET"])
@admin_bp.route("/admin/daily-shift/export", methods=["GET"])
def admin_export_daily_shift_sheet():
    if not session.get("logged_in"):
        return redirect("/login")

    target_d = parse_ymd(request.args.get("target_date") or request.args.get("confirm_date"))
    if not target_d:
        return "\u0074\u0061\u0072\u0067\u0065\u0074\u005f\u0064\u0061\u0074\u0065 \u304c\u4e0d\u6b63\u3067\u3059", 400

    ymd = to_ymd(target_d)
    session["daily_shift_date"] = ymd
    confirmed_rows = get_confirmed_shifts_by_date(ymd)
    active_users = get_all_users(include_inactive=False)
    user_names = [user["name"] or "" for user in active_users if int(user["active"]) == 1 and (user["name"] or "")]
    graph_range = get_daily_shift_graph_range_for_date(ymd, confirmed_rows)
    graph_start_time = graph_range["start_time"]
    graph_end_time = graph_range["end_time"]
    minutes_range = _time_range_minutes(graph_start_time, graph_end_time)
    if not minutes_range:
        return "\u8868\u793a\u958b\u59cb\u6642\u9593 / \u8868\u793a\u7d42\u4e86\u6642\u9593\u304c\u4e0d\u6b63\u3067\u3059", 400
    start_minutes, end_minutes = minutes_range
    if end_minutes - start_minutes > 24 * 60:
        return "\u8868\u793a\u7bc4\u56f2\u306f\u6700\u592724\u6642\u9593\u4ee5\u5185\u306b\u3057\u3066\u304f\u3060\u3055\u3044", 400

    time_options = _build_daily_sheet_time_options(graph_start_time, graph_end_time, 15)
    graph_slots = _build_daily_sheet_time_options(graph_start_time, graph_end_time, 30)[:-1]

    wb = Workbook()
    ws = wb.active
    ws.title = "\u0031\u65e5\u30b7\u30d5\u30c8\u8868"
    list_ws = wb.create_sheet("lists")
    list_ws.sheet_state = "hidden"
    for row_num, name in enumerate(user_names, start=1):
        list_ws.cell(row=row_num, column=1, value=name)
    for row_num, time_label in enumerate(time_options, start=1):
        list_ws.cell(row=row_num, column=2, value=time_label.replace("\u7fcc", ""))

    thin_side = Side(style="thin", color="B8C2CF")
    medium_side = Side(style="medium", color="6B7280")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    hour_border = Border(left=medium_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    input_fill = PatternFill("solid", fgColor="F8FAFC")
    graph_even_fill = PatternFill("solid", fgColor="F3F4F6")
    graph_odd_fill = PatternFill("solid", fgColor="FFFFFF")
    shift_fill = PatternFill("solid", fgColor="00B050")

    ws["A1"] = f"\u0031\u65e5\u30b7\u30d5\u30c8\u8868 {ymd}\uff08{get_weekday_jp(target_d)}\uff09"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A2"] = "\u0057\u0065\u0062\u306e\u0031\u65e5\u30b7\u30d5\u30c8\u8868\u3067\u4fdd\u5b58\u3055\u308c\u305f\u78ba\u5b9a\u30b7\u30d5\u30c8\u3092\u3001\u5370\u5237\u30fb\u5171\u6709\u7528\u306b\u8868\u793a\u3057\u307e\u3059\u3002"

    header_row = 4
    first_data_row = header_row + 1
    extra_blank_rows = 10
    data_row_count = max(len(confirmed_rows) + extra_blank_rows, extra_blank_rows)
    last_data_row = first_data_row + data_row_count - 1
    graph_start_col = 4
    graph_end_col = graph_start_col + len(graph_slots) - 1

    ws.cell(row=header_row, column=1, value="\u540d\u524d")
    ws.cell(row=header_row, column=2, value="\u52e4\u52d9\u958b\u59cb")
    ws.cell(row=header_row, column=3, value="\u52e4\u52d9\u7d42\u4e86")
    for col in range(1, 4):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for index, time_label in enumerate(graph_slots):
        col = graph_start_col + index
        numeric_time = _daily_sheet_label_to_numeric_time(time_label)
        header = ws.cell(row=header_row, column=col, value=numeric_time)
        header.font = header_font
        header.fill = header_fill
        header.alignment = center
        header.border = hour_border if index % 2 == 0 else border
        header.number_format = "hh:mm"

    for index, item in enumerate(confirmed_rows):
        row_num = first_data_row + index
        ws.cell(row=row_num, column=1, value=item["name"] or "")
        ws.cell(row=row_num, column=2, value=item["start_time"] or "")
        ws.cell(row=row_num, column=3, value=item["end_time"] or "")
    if not confirmed_rows:
        ws.cell(row=first_data_row, column=1, value="\u78ba\u5b9a\u30b7\u30d5\u30c8\u306a\u3057")

    for row_num in range(first_data_row, last_data_row + 1):
        for col in range(1, graph_end_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = center if col >= 2 else left
            if col <= 3:
                cell.fill = input_fill
            else:
                slot_index = col - graph_start_col
                hour_band = slot_index // 2
                cell.fill = graph_even_fill if hour_band % 2 == 0 else graph_odd_fill
                if slot_index % 2 == 0:
                    cell.border = hour_border

    for index, item in enumerate(confirmed_rows):
        row_num = first_data_row + index
        clipped = _clip_daily_shift_bar(item["start_time"] or "", item["end_time"] or "", start_minutes, end_minutes)
        if not clipped:
            continue
        clipped_start, clipped_end = clipped
        for slot_index, _slot_label in enumerate(graph_slots):
            slot_start = start_minutes + slot_index * 30
            slot_end = min(slot_start + 30, end_minutes)
            if slot_start < clipped_end and slot_end > clipped_start:
                ws.cell(row=row_num, column=graph_start_col + slot_index).fill = shift_fill

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    for col in range(graph_start_col, graph_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.row_dimensions[header_row].height = 30
    for row_num in range(first_data_row, last_data_row + 1):
        ws.row_dimensions[row_num].height = 24
    ws.freeze_panes = "D5"

    for row in ws.iter_rows(min_row=header_row, max_row=last_data_row, min_col=1, max_col=graph_end_col):
        for cell in row:
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "center", vertical="center")

    if user_names:
        name_validation = DataValidation(
            type="list",
            formula1=f"=lists!$A$1:$A${len(user_names)}",
            allow_blank=True,
        )
        ws.add_data_validation(name_validation)
        name_validation.add(f"A{first_data_row}:A{last_data_row}")

    if time_options:
        time_validation = DataValidation(
            type="list",
            formula1=f"=lists!$B$1:$B${len(time_options)}",
            allow_blank=True,
        )
        ws.add_data_validation(time_validation)
        time_validation.add(f"B{first_data_row}:C{last_data_row}")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"daily_shift_sheet_{ymd}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
